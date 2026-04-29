
import base64
import io
import json
from pathlib import Path
from datetime import date, datetime

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.graphics.shapes import Drawing, String, Rect
from reportlab.graphics.charts.barcharts import HorizontalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.pdfgen import canvas as pdf_canvas


# =========================================================
# QSport Sponsorship Valuation Tool — V10.1 UI FIX
# Client/Internal modes + Scenario persistence + Safe Executive PDF
# =========================================================

st.set_page_config(
    page_title="QSport Sponsorship Valuation Tool",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: visible;}
    [data-testid="stToolbar"] {display:flex !important; visibility:visible !important;}
    .block-container {padding-top: 1.2rem; padding-bottom: 2rem;}
    section[data-testid="stSidebar"] {
        background:#f7f8fb;
        border-right:1px solid #e6e8ef;
    }
    header {visibility: visible !important; display: block !important;}
    .qs-card {background:white; border:1px solid #e8eaf0; border-radius:18px; padding:18px 20px; box-shadow:0 4px 18px rgba(20,30,55,.05); margin-bottom:14px;}
    .qs-title {font-size:30px; font-weight:800; letter-spacing:-.03em; color:#111827; line-height:1.1;}
    .qs-subtitle {color:#667085; font-size:14px; margin-top:6px;}
    .qs-pill {display:inline-block; padding:5px 10px; border-radius:999px; font-size:12px; font-weight:700; border:1px solid #e6e8ef; background:#f9fafb; color:#344054; margin-right:6px;}
    .metric-box {background:white; border:1px solid #e8eaf0; border-radius:18px; padding:18px; box-shadow:0 3px 14px rgba(20,30,55,.04); min-height:112px;}
    .metric-label {color:#667085; font-size:12px; font-weight:700; text-transform:uppercase; letter-spacing:.04em;}
    .metric-value {color:#111827; font-size:26px; font-weight:800; margin-top:6px; line-height:1.05;}
    .metric-note {color:#667085; font-size:12px; margin-top:10px; line-height:1.15;}
    .section-title {color:#111827; font-size:19px; font-weight:800; margin-top:10px; margin-bottom:6px;}
    .small-muted {color:#667085; font-size:13px;}
    </style>
    """,
    unsafe_allow_html=True,
)

try:
    st.set_option("client.showErrorDetails", False)
except Exception:
    pass


ASSET_DIR = Path("assets")
LOGO_MELI = ASSET_DIR / "logo_meli.png"
LOGO_QSPORT = ASSET_DIR / "logo_qsport.png"
BENCHMARK_FILE = Path("benchmark_latam_master_v1.xlsx")
SCENARIO_DIR = Path("scenarios")
SCENARIO_DIR.mkdir(exist_ok=True)

SCORE_LABELS = {"Muy Bajo": 1, "Bajo": 2, "Medio": 3, "Alto": 4, "Muy Alto": 5}
SCORE_FROM_VALUE = {v: k for k, v in SCORE_LABELS.items()}
SCORE_COLS = ["Masividad", "Construcción de Marca", "Potencial de Negocio", "Valor Agregado", "Uso Esperado"]
CLIENT_COLS = ["Asset", "Tipo Visible Cliente", "País", "Marca Interesada"]
EDITOR_COLS = CLIENT_COLS + SCORE_COLS

DEFAULT_MAPPING = pd.DataFrame(
    [
        ["Naming Rights", "Naming Rights", 1.00],
        ["Naming Rights Venue / Arena", "Naming Rights", 1.00],
        ["Naming Rights Tournament", "Naming Rights", 0.95],
        ["Main Sponsor", "Main Sponsor", 1.00],
        ["Sponsor", "Sponsor", 1.00],
        ["Official Partner", "Sponsor", 1.15],
        ["Exclusividad", "Sponsor", 1.20],
        ["Data / CRM", "Sponsor", 1.25],
        ["Digital / Social", "Contenido", 0.90],
        ["PR / Comunicación", "Sponsor", 0.80],
        ["Hospitality Premium", "Hospitality", 1.20],
        ["Hospitality", "Hospitality", 1.00],
        ["Signage / Visibilidad", "Signage", 1.00],
        ["Experiencias VIP", "Hospitality", 1.10],
        ["Contenido", "Contenido", 1.00],
        ["Jersey Sponsor", "Jersey Sponsor", 1.00],
        ["Helmet Sponsor", "Helmet Sponsor", 1.00],
        ["Otros", "Sponsor", 0.70],
    ],
    columns=["Tipo Visible Cliente", "Familia Benchmark", "Multiplicador Subtipo"],
)

DEFAULT_FX = pd.DataFrame([["USD", 1.0], ["BRL", 5.5], ["ARS", 1000.0], ["CLP", 900.0], ["COP", 3900.0], ["EUR", 0.92]], columns=["Moneda", "Local Units per USD"])
DEFAULT_COUNTRY = pd.DataFrame([["LATAM",1.0],["AR",0.9],["BR",1.15],["CL",1.0],["CO",0.85],["MX",1.1]], columns=["País","Factor"])
DEFAULT_INDUSTRY = pd.DataFrame([["Ecommerce / Fintech",1.15],["Betting",1.2],["Insurance",0.95],["Finance",1.05],["Payments",1.1],["Otros",1.0]], columns=["Industria","Factor"])
DEFAULT_BRAND = pd.DataFrame([["Mercado Livre","Ecommerce / Fintech"],["Mercado Libre","Ecommerce / Fintech"],["Meli","Ecommerce / Fintech"],["BetPlay","Betting"],["Bet Play","Betting"],["Betsson","Betting"],["Codere","Betting"],["La Caja","Insurance"],["Allianz","Insurance"],["Visa","Payments"],["Mastercard","Payments"]], columns=["Marca","Industria"])


def b64(path: Path):
    if not path.exists():
        return None
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def show_logo(path: Path, width=180):
    if path.exists():
        st.image(str(path), width=width)


def nt(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def nc(x):
    x = nt(x).upper()
    aliases = {
        "ARG": "AR", "ARGENTINA": "AR",
        "BRA": "BR", "BRASIL": "BR", "BRAZIL": "BR",
        "CHILE": "CL", "COLOMBIA": "CO",
        "MEXICO": "MX", "MÉXICO": "MX",
    }
    return aliases.get(x, x if x else "LATAM")


def clean_price(x):
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = str(x).replace("$", "").replace("USD", "").replace("BRL", "").replace("ARS", "").replace("CLP", "").replace("COP", "").strip()
    s = s.replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def score_to_label(x):
    if isinstance(x, str) and x in SCORE_LABELS:
        return x
    try:
        n = int(round(float(x)))
        n = max(1, min(5, n))
        return SCORE_FROM_VALUE.get(n, "Medio")
    except Exception:
        return "Medio"


def classify(score):
    if score >= 4.5:
        return "Core Premium"
    if score >= 3.5:
        return "Strategic Value"
    if score >= 2.5:
        return "Supporting Asset"
    return "Low Priority"


def money(v):
    try:
        return f"USD {v:,.1f}M"
    except Exception:
        return "USD 0.0M"


def q(series, qq):
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) == 0:
        return 0.0
    return float(s.quantile(qq))


def read_sheet(xls, name, fallback):
    if name in xls.sheet_names:
        return pd.read_excel(xls, sheet_name=name)
    return fallback.copy()


def norm_mapping(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    ren = {}
    for c in df.columns:
        l = c.lower()
        if "visible" in l or "cliente" in l:
            ren[c] = "Tipo Visible Cliente"
        elif "familia" in l or ("benchmark" in l and "scope" not in l):
            ren[c] = "Familia Benchmark"
        elif "multiplicador" in l or "subtipo" in l:
            ren[c] = "Multiplicador Subtipo"
    df = df.rename(columns=ren)
    for col in ["Tipo Visible Cliente", "Familia Benchmark", "Multiplicador Subtipo"]:
        if col not in df.columns:
            return DEFAULT_MAPPING.copy()
    df["Tipo Visible Cliente"] = df["Tipo Visible Cliente"].apply(nt)
    df["Familia Benchmark"] = df["Familia Benchmark"].apply(nt)
    df["Multiplicador Subtipo"] = pd.to_numeric(df["Multiplicador Subtipo"], errors="coerce").fillna(1.0)
    df = df[df["Tipo Visible Cliente"] != ""].copy()
    return df


def norm_fx(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if "Moneda" not in df.columns:
        df.columns = ["Moneda"] + list(df.columns[1:])
    # En tu maestro la columna es Local Units per USD, o sea USD = Precio / FX
    if "Local Units per USD" not in df.columns:
        for c in df.columns:
            if "usd" in c.lower() and "rate" in c.lower():
                df = df.rename(columns={c: "Local Units per USD"})
                break
    if "Local Units per USD" not in df.columns:
        df["Local Units per USD"] = 1.0
    df["Moneda"] = df["Moneda"].apply(lambda x: nt(x).upper())
    df["Local Units per USD"] = pd.to_numeric(df["Local Units per USD"], errors="coerce").fillna(1.0)
    return df


def norm_factor(df, key):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if key not in df.columns:
        df.columns = [key] + list(df.columns[1:])
    if "Factor" not in df.columns:
        df["Factor"] = 1.0
    df[key] = df[key].apply(nc if key == "País" else nt)
    df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce").fillna(1.0)
    return df[[key, "Factor"]].drop_duplicates()


def norm_brand(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if "Marca" not in df.columns:
        df.columns = ["Marca"] + list(df.columns[1:])
    if "Industria" not in df.columns:
        df["Industria"] = "Otros"
    df["Marca"] = df["Marca"].apply(nt)
    df["Industria"] = df["Industria"].apply(nt)
    return df[["Marca", "Industria"]].drop_duplicates()


def infer_ind(brand, brand_map):
    b = nt(brand).lower()
    for _, row in brand_map.iterrows():
        k = nt(row["Marca"]).lower()
        if k and k in b:
            return nt(row["Industria"]) or "Otros"
    return "Otros"


def norm_deals(df, mapping, fx, brand_map):
    if df is None or df.empty:
        return pd.DataFrame(columns=["País","Propiedad","Marca","Año","Tipo Asset Visible","Familia Benchmark","Precio Anual","Moneda","Precio USD Anual","Precio USD Anual M","Industria","Duración","Fuente"])

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    ren = {}
    for c in df.columns:
        l = c.lower().strip()
        if l in ["pais", "país", "country"]:
            ren[c] = "País"
        elif l in ["propiedad", "property"]:
            ren[c] = "Propiedad"
        elif l in ["marca", "brand"]:
            ren[c] = "Marca"
        elif l in ["año", "ano", "year"]:
            ren[c] = "Año"
        elif l in ["tipo asset visible", "tipo visible cliente", "tipo asset", "tipo de asset", "asset type"]:
            ren[c] = "Tipo Asset Visible"
        elif l in ["familia benchmark", "benchmark family"]:
            ren[c] = "Familia Benchmark"
        elif l in ["precio anual", "precio", "valor", "fee", "precio original", "precio original anual"]:
            ren[c] = "Precio Anual"
        elif l in ["moneda", "currency", "moneda original"]:
            ren[c] = "Moneda"
        elif l in ["duración contrato", "duracion contrato", "duración", "duracion", "duration"]:
            ren[c] = "Duración"
        elif l in ["industria", "industry", "industria original"]:
            ren[c] = "Industria"
        elif l in ["fuente", "source"]:
            ren[c] = "Fuente"
    df = df.rename(columns=ren)

    for c in ["País","Propiedad","Marca","Año","Tipo Asset Visible","Familia Benchmark","Precio Anual","Moneda","Duración","Industria","Fuente"]:
        if c not in df.columns:
            df[c] = ""

    df["País"] = df["País"].apply(nc)
    df["Tipo Asset Visible"] = df["Tipo Asset Visible"].apply(nt)
    df["Familia Benchmark"] = df["Familia Benchmark"].apply(nt)
    df["Precio Anual"] = df["Precio Anual"].apply(clean_price)
    df["Moneda"] = df["Moneda"].apply(lambda x: nt(x).upper() or "USD")
    df["Marca"] = df["Marca"].apply(nt)
    df["Industria"] = df.apply(lambda r: nt(r["Industria"]) if nt(r["Industria"]) else infer_ind(r["Marca"], brand_map), axis=1)

    map_dict = dict(zip(mapping["Tipo Visible Cliente"], mapping["Familia Benchmark"]))
    df["Familia Benchmark"] = df.apply(
        lambda r: nt(r["Familia Benchmark"]) if nt(r["Familia Benchmark"]) else map_dict.get(nt(r["Tipo Asset Visible"]), nt(r["Tipo Asset Visible"])),
        axis=1
    )

    fx_dict = dict(zip(fx["Moneda"], fx["Local Units per USD"]))
    df["FX"] = df["Moneda"].map(fx_dict).fillna(1.0)
    # Precio cargado = anual. Local Units per USD: USD = Precio / FX.
    df["Precio USD Anual"] = np.where(df["Moneda"].eq("USD"), df["Precio Anual"], df["Precio Anual"] / df["FX"].replace(0, 1))
    df["Precio USD Anual M"] = df["Precio USD Anual"] / 1_000_000
    df = df[df["Precio USD Anual M"] > 0].copy()
    return df


@st.cache_data(show_spinner=False)
def load_master(file_bytes=None):
    if file_bytes:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
    elif BENCHMARK_FILE.exists():
        xls = pd.ExcelFile(BENCHMARK_FILE)
    else:
        return None

    mapping = norm_mapping(read_sheet(xls, "Asset_Mapping", DEFAULT_MAPPING))
    fx = norm_fx(read_sheet(xls, "FX_Rates", DEFAULT_FX))
    country = norm_factor(read_sheet(xls, "Country_Multipliers", DEFAULT_COUNTRY), "País")
    industry = norm_factor(read_sheet(xls, "Industry_Multipliers", DEFAULT_INDUSTRY), "Industria")
    brand = norm_brand(read_sheet(xls, "Brand_Industry_Map", DEFAULT_BRAND))
    deals_raw = read_sheet(xls, "Deals_Raw", pd.DataFrame())
    deals = norm_deals(deals_raw, mapping, fx, brand)

    return {"mapping": mapping, "fx": fx, "country_mult": country, "industry_mult": industry, "brand_industry": brand, "deals": deals}


def default_client(mapping):
    visible = mapping["Tipo Visible Cliente"].tolist()
    t1 = "Naming Rights Venue / Arena" if "Naming Rights Venue / Arena" in visible else visible[0]
    t2 = "Exclusividad" if "Exclusividad" in visible else visible[min(1, len(visible)-1)]
    t3 = "Hospitality Premium" if "Hospitality Premium" in visible else visible[min(2, len(visible)-1)]
    return pd.DataFrame(
        [
            ["Naming Rights Arena", t1, "BR", "Mercado Livre", "Muy Alto", "Muy Alto", "Alto", "Muy Alto", "Muy Alto"],
            ["Exclusividad categoría", t2, "BR", "Mercado Livre", "Medio", "Muy Alto", "Muy Alto", "Alto", "Muy Alto"],
            ["Hospitality premium", t3, "BR", "Mercado Livre", "Medio", "Alto", "Alto", "Muy Alto", "Alto"],
        ],
        columns=EDITOR_COLS,
    )


def norm_client(df, mapping):
    df = df.copy()
    aliases = {
        "Tipo de Asset": "Tipo Visible Cliente",
        "Tipo Asset": "Tipo Visible Cliente",
        "Marca": "Marca Interesada",
    }
    for old, new in aliases.items():
        if old in df.columns and new not in df.columns:
            df[new] = df[old]

    for c in EDITOR_COLS:
        if c not in df.columns:
            df[c] = "Medio" if c in SCORE_COLS else ("LATAM" if c == "País" else "")

    df = df[EDITOR_COLS].copy()
    df["País"] = df["País"].apply(nc)
    for c in SCORE_COLS:
        df[c] = df[c].apply(score_to_label)

    valid = set(mapping["Tipo Visible Cliente"].tolist())
    df["Tipo Visible Cliente"] = df["Tipo Visible Cliente"].apply(lambda x: nt(x) if nt(x) in valid else "Otros")
    return df


def get_factor(df, key, value):
    if df is None or df.empty or key not in df.columns:
        return 1.0
    rows = df[df[key] == value]
    return float(rows.iloc[0]["Factor"]) if not rows.empty else 1.0


def mapping_row(mapping, visible):
    rows = mapping[mapping["Tipo Visible Cliente"] == visible]
    if rows.empty:
        return "Sponsor", 1.0
    r = rows.iloc[0]
    return nt(r["Familia Benchmark"]), float(r["Multiplicador Subtipo"])


def comps(deals, family, country, industry):
    attempts = [
        (deals[(deals["Familia Benchmark"] == family) & (deals["País"] == country) & (deals["Industria"] == industry)], "Familia + País + Industria"),
        (deals[(deals["Familia Benchmark"] == family) & (deals["País"] == country)], "Familia + País"),
        (deals[(deals["Familia Benchmark"] == family) & (deals["Industria"] == industry)], "Familia + Industria"),
        (deals[deals["Familia Benchmark"] == family], "Familia LATAM"),
        (deals.copy(), "Benchmark general"),
    ]
    for d, method in attempts:
        if len(d) >= 2:
            return d, method
    for d, method in attempts[:-1]:
        if len(d) == 1:
            return d, method + " (1 caso)"
    return attempts[-1]


def calc(client, master):
    rows = []
    for _, r in client.iterrows():
        asset = nt(r["Asset"])
        visible = nt(r["Tipo Visible Cliente"])
        country = nc(r["País"])
        brand = nt(r["Marca Interesada"])
        industry = infer_ind(brand, master["brand_industry"])
        family, sub_mult = mapping_row(master["mapping"], visible)

        score = round(float(np.mean([SCORE_LABELS.get(r[c], 3) for c in SCORE_COLS])), 2)
        score_factor = score / 5.0

        comp, method = comps(master["deals"], family, country, industry)
        low, med, high = q(comp["Precio USD Anual M"], .25), q(comp["Precio USD Anual M"], .5), q(comp["Precio USD Anual M"], .75)

        cf = get_factor(master["country_mult"], "País", country)
        indf = get_factor(master["industry_mult"], "Industria", industry)

        estimated = med * sub_mult * score_factor * cf * indf
        rows.append({
            "Asset": asset,
            "Tipo Visible Cliente": visible,
            "Familia Benchmark": family,
            "País": country,
            "Marca Interesada": brand,
            "Industria": industry,
            "Score Final": score,
            "Clasificación": classify(score),
            "Benchmark Bajo USD M": round(low, 2),
            "Benchmark Medio USD M": round(med, 2),
            "Benchmark Alto USD M": round(high, 2),
            "Multiplicador Subtipo": round(sub_mult, 2),
            "Factor País": round(cf, 2),
            "Factor Industria": round(indf, 2),
            "Valor Estimado USD M": round(estimated, 2),
            "Rango Bajo Ajustado USD M": round(low * sub_mult * score_factor * cf * indf, 2),
            "Rango Alto Ajustado USD M": round(high * sub_mult * score_factor * cf * indf, 2),
            "Comparables": int(len(comp)),
            "Criterio Benchmark": method,
        })
    return pd.DataFrame(rows)


def build_excel(result, client, summary, bench_summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation_Result"
    dark, grey = "111827", "F3F4F6"
    thin = Side(style="thin", color="D0D5DD")

    ws["A1"] = "QSport Sponsorship Valuation Tool"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:N1")
    meta = [
        ("Cliente", summary["client"]),
        ("Proyecto", summary["project"]),
        ("Precio solicitado USD M", summary["asking_price"]),
        ("Valor estimado USD M", summary["estimated_value"]),
        ("Gap USD M", summary["gap"]),
        ("Gap %", summary["gap_pct"]),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor=grey)
        ws.cell(i, 2, v)

    cols = list(result.columns)
    start = 11
    for j, col in enumerate(cols, 1):
        c = ws.cell(start, j, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for i, (_, row) in enumerate(result.iterrows(), start + 1):
        for j, col in enumerate(cols, 1):
            c = ws.cell(i, j, row[col])
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for idx in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 22

    for name, data in [("Client_Input", client), ("Benchmark_Summary", bench_summary)]:
        s = wb.create_sheet(name)
        for j, col in enumerate(data.columns, 1):
            s.cell(1, j, col).font = Font(bold=True)
        for i, (_, row) in enumerate(data.iterrows(), 2):
            for j, col in enumerate(data.columns, 1):
                s.cell(i, j, row[col])

    m = wb.create_sheet("Methodology")
    items = [
        ["Principio", "Estimación de valor con benchmarks internos QSport y scoring estratégico."],
        ["Fórmula", "Valor = Benchmark mediano x multiplicador subtipo x score/5 x factor país x factor industria."],
        ["Benchmark", "Prioriza Familia + País + Industria, luego Familia + País, Familia + Industria y Familia LATAM."],
        ["Precio", "Los precios cargados en Deals_Raw son tratados como valores anuales."],
    ]
    for i, row in enumerate(items, 1):
        m.cell(i, 1, row[0]).font = Font(bold=True)
        m.cell(i, 2, row[1])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _pdf_safe(v):
    return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _pdf_money(v):
    try:
        return f"USD {float(v):,.2f}M"
    except Exception:
        return "USD 0.00M"


def _short_text(v, n=34):
    s = str(v)
    return s if len(s) <= n else s[: n - 1] + "…"


def _brand_cover(canvas, doc, summary, logo_qsport, logo_client):
    canvas.saveState()
    w, h = landscape(A4)
    canvas.setFillColor(colors.HexColor("#0B1220"))
    canvas.rect(0, 0, w, h, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#F2C94C"))
    canvas.rect(0, h - 0.45 * cm, w, 0.45 * cm, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 29)
    canvas.drawString(1.55 * cm, h - 3.0 * cm, "Sponsorship / Naming Rights")
    canvas.drawString(1.55 * cm, h - 4.0 * cm, "Valuation Report")
    canvas.setFillColor(colors.HexColor("#D0D5DD"))
    canvas.setFont("Helvetica", 12)
    canvas.drawString(1.6 * cm, h - 5.05 * cm, "Independent market valuation based on QSport benchmark intelligence")
    canvas.setStrokeColor(colors.HexColor("#344054"))
    canvas.line(1.6 * cm, h - 5.65 * cm, 13.2 * cm, h - 5.65 * cm)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawString(1.6 * cm, 4.0 * cm, f"Prepared for: {summary.get('client','')}")
    canvas.setFillColor(colors.HexColor("#D0D5DD"))
    canvas.setFont("Helvetica", 11)
    canvas.drawString(1.6 * cm, 3.25 * cm, f"Project: {summary.get('project','')}")
    canvas.drawString(1.6 * cm, 2.55 * cm, f"Date: {date.today().isoformat()}")
    for path, x, y in [(logo_qsport, w - 6.4 * cm, h - 3.05 * cm), (logo_client, w - 6.4 * cm, 2.4 * cm)]:
        try:
            if path and Path(path).exists():
                canvas.drawImage(str(path), x, y, width=4.8*cm, height=1.6*cm, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    canvas.setFillColor(colors.HexColor("#F2C94C"))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(1.6 * cm, 1.35 * cm, "CONFIDENTIAL EXECUTIVE OUTPUT")
    canvas.restoreState()


def _page_footer(canvas, doc, summary):
    canvas.saveState()
    w, h = landscape(A4)
    canvas.setStrokeColor(colors.HexColor("#EAECF0"))
    canvas.line(1.2 * cm, 0.8 * cm, w - 1.2 * cm, 0.8 * cm)
    canvas.setFillColor(colors.HexColor("#98A2B3"))
    canvas.setFont("Helvetica", 7)
    canvas.drawString(1.2 * cm, 0.55 * cm, "QSport Sponsorship Valuation Tool | Confidential")
    canvas.drawRightString(w - 1.2 * cm, 0.55 * cm, f"{summary.get('client','')} - {summary.get('project','')} | Page {doc.page}")
    canvas.restoreState()


def _kpi_card(title, value, note="", accent="#111827"):
    styles = getSampleStyleSheet()
    tbl = Table([
        [Paragraph(f'<font color="#667085" size="7"><b>{_pdf_safe(title).upper()}</b></font>', styles["Normal"])],
        [Paragraph(f'<font color="{accent}" size="16"><b>{_pdf_safe(value)}</b></font>', styles["Normal"])],
        [Paragraph(f'<font color="#667085" size="7">{_pdf_safe(note)}</font>', styles["Normal"])]
    ], colWidths=[6.05*cm], rowHeights=[0.55*cm, 0.9*cm, 0.45*cm])
    tbl.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.white), ("BOX", (0,0), (-1,-1), .8, colors.HexColor("#EAECF0")), ("LEFTPADDING", (0,0), (-1,-1), 12), ("RIGHTPADDING", (0,0), (-1,-1), 12), ("TOPPADDING", (0,0), (-1,-1), 6), ("BOTTOMPADDING", (0,0), (-1,-1), 4)]))
    return tbl


def bar_draw(labels, values, width=700, height=260):
    d = Drawing(width, height)
    maxv = max(values) if values else 1
    maxv = maxv if maxv > 0 else 1
    left = 185
    top = height - 24
    bar_h = 17
    gap = 9
    chart_w = width - left - 86
    for i, (lab, val) in enumerate(zip(labels, values)):
        y = top - i * (bar_h + gap) - bar_h
        d.add(String(0, y + 4, _short_text(lab, 31), fontName="Helvetica", fontSize=8, fillColor=colors.HexColor("#344054")))
        d.add(Rect(left, y, chart_w, bar_h, fillColor=colors.HexColor("#F2F4F7"), strokeColor=None))
        bw = chart_w * (float(val) / maxv)
        d.add(Rect(left, y, bw, bar_h, fillColor=colors.HexColor("#111827"), strokeColor=None))
        d.add(String(left + bw + 5, y + 4, f"{float(val):.2f}", fontName="Helvetica-Bold", fontSize=8, fillColor=colors.HexColor("#111827")))
    return d


def share_table_draw(df):
    styles = getSampleStyleSheet()
    total = float(df["Valor Estimado USD M"].sum()) if not df.empty else 0.0
    data = [["Asset type", "Value", "Share"]]
    for _, r in df.head(8).iterrows():
        v = float(r["Valor Estimado USD M"])
        sh = v / total * 100 if total else 0
        bar = "█" * max(1, int(sh / 5))
        data.append([Paragraph(_pdf_safe(_short_text(r["Tipo Visible Cliente"], 26)), styles["Normal"]), f"{v:.2f}", f"{sh:.1f}%  {bar}"])
    tbl = Table(data, colWidths=[8.4*cm, 3.0*cm, 6.0*cm])
    tbl.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 8), ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#EAECF0")), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]), ("ALIGN", (1,1), (-1,-1), "RIGHT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("LEFTPADDING", (0,0), (-1,-1), 8), ("RIGHTPADDING", (0,0), (-1,-1), 8)]))
    return tbl


def _narrative_summary(summary):
    gap_pct = float(summary.get("gap_pct", 0))
    if gap_pct <= -10:
        return "The requested price is materially above the estimated benchmark value. The current package should be renegotiated, reinforced with additional high-value rights, or repositioned before approval."
    if gap_pct >= 10:
        return "The estimated benchmark value is above the requested price. The package presents a favorable acquisition opportunity if contractual rights and exclusivity are secured."
    return "The requested price is broadly aligned with the estimated benchmark range. The decision should focus on rights quality, category exclusivity and activation capacity."


def build_pdf(result, summary, logo_qsport, logo_client):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1.15*cm, leftMargin=1.15*cm, topMargin=1.0*cm, bottomMargin=1.0*cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="QTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=22, leading=25, textColor=colors.HexColor("#111827"), alignment=0, spaceAfter=8))
    styles.add(ParagraphStyle(name="QH2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=colors.HexColor("#111827"), spaceBefore=8, spaceAfter=6))
    styles.add(ParagraphStyle(name="QBody", parent=styles["Normal"], fontSize=9, leading=12, textColor=colors.HexColor("#475467")))
    styles.add(ParagraphStyle(name="QSmall", parent=styles["Normal"], fontSize=7.2, leading=9, textColor=colors.HexColor("#667085")))
    styles.add(ParagraphStyle(name="QCell", parent=styles["Normal"], fontSize=7.4, leading=9, textColor=colors.HexColor("#111827")))
    story = [Spacer(1, 0.1 * cm), PageBreak()]
    story.append(Paragraph("Executive Summary", styles["QTitle"]))
    accent = "#D92D20" if float(summary.get("gap_pct", 0)) < -10 else ("#027A48" if float(summary.get("gap_pct", 0)) > 10 else "#111827")
    cards_tbl = Table([[
        _kpi_card("Estimated value", _pdf_money(summary["estimated_value"]), "Benchmark + scoring"),
        _kpi_card("Asking price", _pdf_money(summary["asking_price"]), "Property requested price"),
        _kpi_card("Gap", _pdf_money(summary["gap"]), f"{summary['gap_pct']:.1f}% vs asking price", accent),
        _kpi_card("Assets / comparables", f"{len(result)} / {int(result['Comparables'].sum()) if not result.empty else 0}", "Valued rights and cases used"),
    ]], colWidths=[6.35*cm]*4, hAlign="LEFT")
    cards_tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 0), ("RIGHTPADDING", (0,0), (-1,-1), 8)]))
    rec_box = Table([[Paragraph(f'<font color="#111827"><b>Strategic reading</b></font><br/>{_pdf_safe(_narrative_summary(summary))}', styles["QBody"])]], colWidths=[26.0*cm])
    rec_box.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F9FAFB")), ("BOX", (0,0), (-1,-1), .8, colors.HexColor("#EAECF0")), ("LEFTPADDING", (0,0), (-1,-1), 12), ("RIGHTPADDING", (0,0), (-1,-1), 12), ("TOPPADDING", (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10)]))
    story += [cards_tbl, Spacer(1, .45*cm), rec_box, Spacer(1, .42*cm), Paragraph("Top valued assets", styles["QH2"])]
    top = result.sort_values("Valor Estimado USD M", ascending=False).head(10)
    table = [["Asset", "Type", "Score", "Class", "Value", "Range", "Comp."]]
    for _, r in top.iterrows():
        table.append([Paragraph(_pdf_safe(_short_text(r["Asset"], 44)), styles["QCell"]), Paragraph(_pdf_safe(_short_text(r["Tipo Visible Cliente"], 30)), styles["QCell"]), f"{float(r['Score Final']):.2f}", Paragraph(_pdf_safe(str(r["Clasificación"])), styles["QCell"]), f"{float(r['Valor Estimado USD M']):.2f}", f"{float(r['Rango Bajo Ajustado USD M']):.2f} - {float(r['Rango Alto Ajustado USD M']):.2f}", str(int(r["Comparables"]))])
    tt = Table(table, colWidths=[7.2*cm, 5.0*cm, 1.7*cm, 3.0*cm, 2.2*cm, 3.0*cm, 1.5*cm], repeatRows=1)
    tt.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 7.5), ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#D0D5DD")), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]), ("ALIGN", (2,1), (-1,-1), "RIGHT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("LEFTPADDING", (0,0), (-1,-1), 6), ("RIGHTPADDING", (0,0), (-1,-1), 6)]))
    story += [tt, PageBreak(), Paragraph("Value Architecture", styles["QTitle"])]
    ctop = result.sort_values("Valor Estimado USD M", ascending=False).head(10)
    story += [Paragraph("Estimated value by asset", styles["QH2"]), bar_draw([str(x) for x in ctop["Asset"].tolist()], [float(x) for x in ctop["Valor Estimado USD M"].tolist()]), Spacer(1, .35*cm)]
    type_df = result.groupby("Tipo Visible Cliente", as_index=False)["Valor Estimado USD M"].sum().sort_values("Valor Estimado USD M", ascending=False)
    left = share_table_draw(type_df)
    concentration = float(type_df.iloc[0]["Valor Estimado USD M"] / type_df["Valor Estimado USD M"].sum() * 100) if not type_df.empty and type_df["Valor Estimado USD M"].sum() else 0
    right = Table([[Paragraph(f'<font color="#111827"><b>Portfolio concentration</b></font><br/><br/>The leading rights family represents <b>{concentration:.1f}%</b> of the estimated package value. This helps identify whether valuation is driven by one premium asset or by a balanced rights portfolio.<br/><br/><font color="#667085">Use this page to stress-test negotiation priorities and rights that need reinforcement.</font>', styles["QBody"])]], colWidths=[7.6*cm])
    right.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F9FAFB")), ("BOX", (0,0), (-1,-1), .8, colors.HexColor("#EAECF0")), ("LEFTPADDING", (0,0), (-1,-1), 12), ("RIGHTPADDING", (0,0), (-1,-1), 12), ("TOPPADDING", (0,0), (-1,-1), 12)]))
    story += [Paragraph("Distribution by asset type", styles["QH2"]), Table([[left, right]], colWidths=[18.0*cm, 7.9*cm], style=[("VALIGN", (0,0), (-1,-1), "TOP")]), PageBreak()]
    story += [Paragraph("Asset Detail", styles["QTitle"])]
    detail = [["Asset", "Benchmark family", "Country", "Industry", "Score", "Value", "Benchmark criteria"]]
    for _, r in result.sort_values("Valor Estimado USD M", ascending=False).iterrows():
        detail.append([Paragraph(_pdf_safe(_short_text(r["Asset"], 38)), styles["QCell"]), Paragraph(_pdf_safe(_short_text(r["Familia Benchmark"], 24)), styles["QCell"]), str(r["País"]), Paragraph(_pdf_safe(_short_text(r["Industria"], 22)), styles["QCell"]), f"{float(r['Score Final']):.2f}", f"{float(r['Valor Estimado USD M']):.2f}", Paragraph(_pdf_safe(_short_text(r["Criterio Benchmark"], 34)), styles["QCell"])])
    dt = Table(detail, colWidths=[6.6*cm, 4.2*cm, 1.5*cm, 3.7*cm, 1.6*cm, 2.0*cm, 5.1*cm], repeatRows=1)
    dt.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 7.1), ("GRID", (0,0), (-1,-1), .3, colors.HexColor("#D0D5DD")), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]), ("ALIGN", (2,1), (2,-1), "CENTER"), ("ALIGN", (4,1), (5,-1), "RIGHT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("LEFTPADDING", (0,0), (-1,-1), 5), ("RIGHTPADDING", (0,0), (-1,-1), 5)]))
    story += [dt, PageBreak()]
    story += [Paragraph("Methodology", styles["QTitle"])]
    methodology = [["1", "Benchmark selection", "The engine prioritizes comparables by family, country and industry. If the sample is insufficient, it progressively broadens the benchmark scope."], ["2", "Strategic scoring", "Each right is rated across masividad, brand building, business potential, added value and expected use. The final score adjusts the benchmark median."], ["3", "Commercial adjustment", "Subtype, country and industry factors are applied to better reflect the local market and the value of each right."], ["4", "Independence", "The valuation estimates market value independently from the asking price. The full benchmark database remains proprietary."]]
    mt = Table([[Paragraph(f'<font color="#111827"><b>{a}. {b}</b></font><br/>{_pdf_safe(c)}', styles["QBody"])] for a,b,c in methodology], colWidths=[25.8*cm])
    mt.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.white), ("BOX", (0,0), (-1,-1), .8, colors.HexColor("#EAECF0")), ("INNERGRID", (0,0), (-1,-1), .35, colors.HexColor("#EAECF0")), ("LEFTPADDING", (0,0), (-1,-1), 12), ("RIGHTPADDING", (0,0), (-1,-1), 12), ("TOPPADDING", (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10)]))
    story += [mt, Spacer(1, .35*cm), Paragraph("Formula", styles["QH2"]), Paragraph("Estimated Value = Median Benchmark x Subtype Multiplier x Score/5 x Country Factor x Industry Factor", styles["QBody"]), Spacer(1, .3*cm), Paragraph("Note: Benchmark prices are treated as annual values. This PDF is designed as a client-facing executive output; detailed benchmark rows are intentionally excluded.", styles["QSmall"])]
    doc.build(story, onFirstPage=lambda canvas, doc: _brand_cover(canvas, doc, summary, logo_qsport, logo_client), onLaterPages=lambda canvas, doc: _page_footer(canvas, doc, summary))
    buf.seek(0)
    return buf.read()


# =========================================================
# V9.4 - Dual export + safe PDF engine
# =========================================================

INTERNAL_PASSWORD_DEFAULT = "qsport"


def _safe_filename(x):
    s = nt(x).replace("/", "-").replace("\\", "-").replace(":", "-")
    return "_".join(s.split()) or "valuation"


def build_client_excel(result, summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Client_Summary"
    dark, grey, light = "111827", "F9FAFB", "EAECF0"
    thin = Side(style="thin", color="D0D5DD")
    ws["A1"] = "QSport Sponsorship Valuation - Client Summary"
    ws["A1"].font = Font(bold=True, size=16, color=dark)
    ws.merge_cells("A1:H1")
    meta = [
        ("Cliente", summary.get("client", "")),
        ("Proyecto", summary.get("project", "")),
        ("Valor estimado USD M", round(float(summary.get("estimated_value", 0)), 2)),
        ("Precio solicitado USD M", round(float(summary.get("asking_price", 0)), 2)),
        ("Gap USD M", round(float(summary.get("gap", 0)), 2)),
        ("Gap %", f"{float(summary.get('gap_pct', 0)):.1f}%"),
        ("Recomendación", summary.get("recommendation", "")),
    ]
    for i, (k, v) in enumerate(meta, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor=grey)
        ws.cell(i, 2, v)
    client_cols = ["Asset", "Tipo Visible Cliente", "Score Final", "Clasificación", "Valor Estimado USD M", "Rango Bajo Ajustado USD M", "Rango Alto Ajustado USD M"]
    start = 13
    for j, col in enumerate(client_cols, 1):
        c = ws.cell(start, j, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=dark)
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c.alignment = Alignment(horizontal="center")
    safe_result = result[client_cols].copy() if not result.empty else pd.DataFrame(columns=client_cols)
    for i, (_, row) in enumerate(safe_result.iterrows(), start + 1):
        for j, col in enumerate(client_cols, 1):
            c = ws.cell(i, j, row[col])
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "USD M" in col or col == "Score Final":
                c.number_format = "0.00"
    notes_start = start + len(safe_result) + 4
    ws.cell(notes_start, 1, "Nota metodológica").font = Font(bold=True)
    ws.cell(notes_start + 1, 1, "Este archivo es un output ejecutivo para cliente. No incluye benchmark interno, comparables, factores ni criterios técnicos propietarios.")
    ws.merge_cells(start_row=notes_start + 1, start_column=1, end_row=notes_start + 1, end_column=7)
    for idx, width in enumerate([34, 28, 14, 20, 20, 24, 24], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _draw_logo(c, path, x, y, w, h):
    try:
        if path and Path(path).exists():
            c.drawImage(str(path), x, y, width=w, height=h, preserveAspectRatio=True, mask="auto")
    except Exception:
        pass


def _draw_header(c, summary, page_title, page_num, logo_qsport=None):
    w, h = landscape(A4)
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(1.2*cm, h-0.75*cm, "QSport Sponsorship Valuation Tool")
    c.setFillColor(colors.HexColor("#98A2B3"))
    c.setFont("Helvetica", 8)
    c.drawRightString(w-1.2*cm, h-0.75*cm, f"{summary.get('client','')} - {summary.get('project','')} | Page {page_num}")
    c.setStrokeColor(colors.HexColor("#EAECF0"))
    c.line(1.2*cm, h-0.95*cm, w-1.2*cm, h-0.95*cm)
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 21)
    c.drawString(1.2*cm, h-1.75*cm, page_title)


def _draw_footer(c, summary):
    w, h = landscape(A4)
    c.setStrokeColor(colors.HexColor("#EAECF0"))
    c.line(1.2*cm, 0.8*cm, w-1.2*cm, 0.8*cm)
    c.setFillColor(colors.HexColor("#98A2B3"))
    c.setFont("Helvetica", 7)
    c.drawString(1.2*cm, 0.52*cm, "Confidential executive output. Benchmark database is proprietary and not disclosed in full.")


def _wrap_lines(text, max_chars):
    words = _pdf_safe(text).split()
    lines, cur = [], ""
    for word in words:
        if len(cur) + len(word) + 1 <= max_chars:
            cur = (cur + " " + word).strip()
        else:
            if cur:
                lines.append(cur)
            cur = word
    if cur:
        lines.append(cur)
    return lines


def _draw_kpi(c, x, y, w_box, h_box, label, value, note, color="#111827"):
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor("#E4E7EC"))
    c.setLineWidth(0.8)
    c.roundRect(x, y, w_box, h_box, 10, fill=1, stroke=1)

    c.setFillColor(colors.HexColor("#667085"))
    c.setFont("Helvetica-Bold", 7.8)
    c.drawString(x + 0.32*cm, y + h_box - 0.55*cm, str(label).upper())

    c.setFillColor(colors.HexColor(color))
    c.setFont("Helvetica-Bold", 12.8)
    c.drawString(x + 0.32*cm, y + 0.74*cm, str(value))

    c.setFillColor(colors.HexColor("#667085"))
    c.setFont("Helvetica", 6.6)
    c.drawString(x + 0.32*cm, y + 0.24*cm, str(note))


def _draw_table(c, x, y, col_widths, headers, rows, row_h=0.58*cm, font_size=7.2, max_rows=None):
    if max_rows is not None:
        rows = rows[:max_rows]
    total_w = sum(col_widths)
    c.setFillColor(colors.HexColor("#111827"))
    c.rect(x, y-row_h, total_w, row_h, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", font_size)
    cx = x
    for i, htxt in enumerate(headers):
        c.drawString(cx+0.12*cm, y-row_h+0.2*cm, str(htxt))
        cx += col_widths[i]
    c.setStrokeColor(colors.HexColor("#D0D5DD"))
    c.rect(x, y-row_h, total_w, row_h, fill=0, stroke=1)
    yy = y-row_h
    for ridx, row in enumerate(rows):
        yy -= row_h
        c.setFillColor(colors.white if ridx % 2 == 0 else colors.HexColor("#F9FAFB"))
        c.rect(x, yy, total_w, row_h, fill=1, stroke=0)
        c.setStrokeColor(colors.HexColor("#EAECF0"))
        c.rect(x, yy, total_w, row_h, fill=0, stroke=1)
        c.setFillColor(colors.HexColor("#111827"))
        c.setFont("Helvetica", font_size)
        cx = x
        for i, val in enumerate(row):
            txt = _short_text(val, max(10, int(col_widths[i]/cm*4.2)))
            c.drawString(cx+0.12*cm, yy+0.2*cm, str(txt))
            cx += col_widths[i]
    return yy


def build_client_pdf(result, summary, logo_qsport, logo_client):
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=landscape(A4))
    w, h = landscape(A4)
    # Cover
    c.setFillColor(colors.HexColor("#101828"))
    c.rect(0, 0, w, h, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 31)
    c.drawString(1.5*cm, h-3.0*cm, "Sponsorship / Naming Rights")
    c.drawString(1.5*cm, h-4.05*cm, "Valuation Report")
    c.setFillColor(colors.HexColor("#D0D5DD"))
    c.setFont("Helvetica", 12)
    c.drawString(1.55*cm, h-4.8*cm, "Independent market valuation based on QSport benchmark intelligence")
    c.setStrokeColor(colors.HexColor("#F2C94C"))
    c.setLineWidth(2)
    c.line(1.55*cm, h-5.35*cm, 12.6*cm, h-5.35*cm)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(1.55*cm, 4.25*cm, f"Prepared for: {summary.get('client','')}")
    c.setFillColor(colors.HexColor("#D0D5DD"))
    c.setFont("Helvetica", 11)
    c.drawString(1.55*cm, 3.48*cm, f"Project: {summary.get('project','')}")
    c.drawString(1.55*cm, 2.78*cm, f"Date: {date.today().isoformat()}")
    c.setFillColor(colors.HexColor("#F2C94C"))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(1.55*cm, 1.45*cm, "CONFIDENTIAL CLIENT OUTPUT")
    _draw_logo(c, logo_qsport, w-6.3*cm, h-3.2*cm, 4.8*cm, 1.7*cm)
    _draw_logo(c, logo_client, w-6.3*cm, 2.35*cm, 4.8*cm, 1.7*cm)
    c.showPage()
    # Executive summary
    _draw_header(c, summary, "Executive Summary", 2, logo_qsport)
    y = h-3.25*cm
    accent = "#D92D20" if float(summary.get("gap_pct", 0)) < -10 else ("#027A48" if float(summary.get("gap_pct", 0)) > 10 else "#111827")
    card_w, gap_x = 6.05*cm, 0.55*cm
    card_h = 2.18*cm
    x0 = 1.2*cm
    _draw_kpi(c, x0, y-card_h, card_w, card_h, "Estimated value", _pdf_money(summary.get("estimated_value", 0)), "Benchmark + scoring")
    _draw_kpi(c, x0+card_w+gap_x, y-card_h, card_w, card_h, "Asking price", _pdf_money(summary.get("asking_price", 0)), "Property requested price")
    _draw_kpi(c, x0+2*(card_w+gap_x), y-card_h, card_w, card_h, "Gap", _pdf_money(summary.get("gap", 0)), f"{float(summary.get('gap_pct', 0)):.1f}% vs asking", accent)
    _draw_kpi(c, x0+3*(card_w+gap_x), y-card_h, card_w, card_h, "Assets", str(len(result)), "Valued rights")
    y -= 3.05*cm
    c.setFillColor(colors.HexColor("#F9FAFB"))
    c.setStrokeColor(colors.HexColor("#EAECF0"))
    c.roundRect(1.2*cm, y-1.55*cm, w-2.4*cm, 1.3*cm, 8, fill=1, stroke=1)
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.55*cm, y-0.65*cm, "Strategic reading")
    c.setFont("Helvetica", 8.5)
    c.setFillColor(colors.HexColor("#475467"))
    yy = y-1.03*cm
    for line in _wrap_lines(_narrative_summary(summary), 146)[:2]:
        c.drawString(1.55*cm, yy, line)
        yy -= 0.34*cm
    y -= 2.05*cm
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(1.2*cm, y, "Top valued assets")
    top = result.sort_values("Valor Estimado USD M", ascending=False).head(10)
    rows = []
    for _, r in top.iterrows():
        rows.append([r["Asset"], r["Tipo Visible Cliente"], f"{float(r['Score Final']):.2f}", r["Clasificación"], f"{float(r['Valor Estimado USD M']):.2f}", f"{float(r['Rango Bajo Ajustado USD M']):.2f} - {float(r['Rango Alto Ajustado USD M']):.2f}"])
    _draw_table(c, 1.2*cm, y-0.45*cm, [7.7*cm, 5.2*cm, 1.8*cm, 3.5*cm, 2.3*cm, 3.5*cm], ["Asset", "Type", "Score", "Class", "Value", "Range"], rows, row_h=0.57*cm, font_size=6.9)
    _draw_footer(c, summary)
    c.showPage()
    # Value architecture
    _draw_header(c, summary, "Value Architecture", 3, logo_qsport)
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(1.2*cm, h-2.6*cm, "Estimated value by asset")
    top_bar = result.sort_values("Valor Estimado USD M", ascending=False).head(10)
    maxv = max([float(x) for x in top_bar["Valor Estimado USD M"]], default=1)
    y = h-3.15*cm
    for _, r in top_bar.iterrows():
        val = float(r["Valor Estimado USD M"])
        c.setFillColor(colors.HexColor("#344054"))
        c.setFont("Helvetica", 8)
        c.drawString(1.2*cm, y, _short_text(r["Asset"], 34))
        c.setFillColor(colors.HexColor("#F2F4F7"))
        c.rect(7.5*cm, y-0.09*cm, 11.5*cm, 0.24*cm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#111827"))
        c.rect(7.5*cm, y-0.09*cm, 11.5*cm*(val/maxv if maxv else 0), 0.24*cm, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(19.3*cm, y, f"{val:.2f}")
        y -= 0.62*cm
    type_df = result.groupby("Tipo Visible Cliente", as_index=False)["Valor Estimado USD M"].sum().sort_values("Valor Estimado USD M", ascending=False)
    total = float(type_df["Valor Estimado USD M"].sum()) if not type_df.empty else 0
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(1.2*cm, 6.7*cm, "Distribution by asset type")
    rows = []
    for _, r in type_df.head(8).iterrows():
        sh = float(r["Valor Estimado USD M"])/total*100 if total else 0
        rows.append([r["Tipo Visible Cliente"], f"{float(r['Valor Estimado USD M']):.2f}", f"{sh:.1f}%"])
    _draw_table(c, 1.2*cm, 6.25*cm, [8.0*cm, 3.0*cm, 3.0*cm], ["Asset type", "Value", "Share"], rows, row_h=0.55*cm, font_size=7.4)
    c.setFillColor(colors.HexColor("#F9FAFB"))
    c.setStrokeColor(colors.HexColor("#EAECF0"))
    c.roundRect(17.2*cm, 2.15*cm, 10.9*cm, 4.6*cm, 8, fill=1, stroke=1)
    concentration = float(type_df.iloc[0]["Valor Estimado USD M"] / total * 100) if total else 0
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(17.65*cm, 5.95*cm, "Portfolio concentration")
    c.setFillColor(colors.HexColor("#475467"))
    c.setFont("Helvetica", 8.5)
    text = f"The leading rights family represents {concentration:.1f}% of the estimated package value. This helps identify whether valuation is driven by one premium asset or by a balanced rights portfolio."
    yy = 5.45*cm
    for line in _wrap_lines(text, 54)[:6]:
        c.drawString(17.65*cm, yy, line)
        yy -= 0.37*cm
    _draw_footer(c, summary)
    c.showPage()
    # Methodology
    _draw_header(c, summary, "Methodology", 4, logo_qsport)
    methods = [
        ("1. Benchmark selection", "The engine prioritizes comparable assets by family, country and industry. If the sample is insufficient, it progressively broadens the benchmark scope."),
        ("2. Strategic scoring", "Each right is rated across masividad, brand building, business potential, added value and expected use. The final score adjusts the benchmark median."),
        ("3. Commercial adjustment", "Subtype, country and industry factors are applied to reflect local market conditions and the relative value of each right."),
        ("4. Independence", "The valuation estimates market value independently from the asking price. Detailed benchmark rows remain proprietary."),
    ]
    y = h-3.0*cm
    for title, body in methods:
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor("#EAECF0"))
        c.roundRect(1.2*cm, y-1.4*cm, w-2.4*cm, 1.1*cm, 7, fill=1, stroke=1)
        c.setFillColor(colors.HexColor("#111827"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(1.55*cm, y-0.72*cm, title)
        c.setFillColor(colors.HexColor("#475467"))
        c.setFont("Helvetica", 8.5)
        c.drawString(7.2*cm, y-0.72*cm, _short_text(body, 138))
        y -= 1.55*cm
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(1.2*cm, y-0.25*cm, "Formula")
    c.setFillColor(colors.HexColor("#475467"))
    c.setFont("Helvetica", 9)
    c.drawString(1.2*cm, y-0.8*cm, "Estimated Value = Median Benchmark x Subtype Multiplier x Score/5 x Country Factor x Industry Factor")
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#667085"))
    c.drawString(1.2*cm, y-1.45*cm, "Note: Benchmark prices are treated as annual values. This PDF is designed as a client-facing executive output.")
    _draw_footer(c, summary)
    c.save()
    buf.seek(0)
    return buf.read()


def build_reloadable_input_excel(client_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Client_Input"
    ws["A1"] = "QSport Sponsorship Valuation - Reloadable Input"
    ws["A1"].font = Font(bold=True, size=15, color="111827")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EDITOR_COLS))
    ws["A2"] = "Archivo recargable: el cliente puede guardar su evaluación y volver a subirlo en la app."
    ws["A2"].font = Font(size=10, color="667085")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(EDITOR_COLS))

    header_row = 4
    dark = "111827"
    grey = "F3F4F6"
    thin = Side(style="thin", color="D0D5DD")
    for j, col in enumerate(EDITOR_COLS, 1):
        cell = ws.cell(header_row, j, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    data = client_df[EDITOR_COLS].copy() if not client_df.empty else pd.DataFrame(columns=EDITOR_COLS)
    for i, (_, row) in enumerate(data.iterrows(), header_row + 1):
        for j, col in enumerate(EDITOR_COLS, 1):
            cell = ws.cell(i, j, row[col])
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if col in SCORE_COLS:
                cell.fill = PatternFill("solid", fgColor=grey)

    widths = {"A": 34, "B": 28, "C": 12, "D": 22, "E": 18, "F": 24, "G": 22, "H": 18, "I": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"

    info = wb.create_sheet("Instrucciones")
    info["A1"] = "Cómo usar este archivo"
    info["A1"].font = Font(bold=True, size=14)
    notes = [
        "1. Completá o editá los assets en la hoja Client_Input.",
        "2. Usá las etiquetas Muy Bajo, Bajo, Medio, Alto o Muy Alto en los campos de scoring.",
        "3. Guardá el archivo y volvelo a cargar en la app desde la sidebar.",
        "4. No cambies los nombres de las columnas para que la app lo lea correctamente.",
    ]
    for idx, note in enumerate(notes, 3):
        info.cell(idx, 1, note)
        info.cell(idx, 1).font = Font(size=10, color="475467")
    info.column_dimensions["A"].width = 95

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def build_internal_pdf(result, summary, logo_qsport, logo_client):
    # Reuse the safe client PDF for first pages and keep the detailed information in the internal Excel.
    return build_client_pdf(result, summary, logo_qsport, logo_client)




# =========================================================
# V10 - Scenario Engine
# =========================================================

def scenario_key(client, project, scenario):
    raw = f"{_safe_filename(client)}__{_safe_filename(project)}__{_safe_filename(scenario)}".strip("_")
    return raw or "scenario"


def scenario_path(client, project, scenario):
    return SCENARIO_DIR / f"{scenario_key(client, project, scenario)}.json"


def scenario_payload(client, project, asking_price, scenario_name, client_df):
    return {
        "version": "V10.3 Scenario Engine",
        "scenario_name": nt(scenario_name) or "Base",
        "client": nt(client),
        "project": nt(project),
        "asking_price": float(asking_price or 0),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "client_input": client_df[EDITOR_COLS].fillna("").to_dict(orient="records") if isinstance(client_df, pd.DataFrame) else [],
    }


def scenario_to_bytes(client, project, asking_price, scenario_name, client_df):
    payload = scenario_payload(client, project, asking_price, scenario_name, client_df)
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def save_scenario_local(client, project, asking_price, scenario_name, client_df):
    payload = scenario_payload(client, project, asking_price, scenario_name, client_df)
    path = scenario_path(client, project, scenario_name)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def load_scenario_payload(payload, mapping):
    # Streamlit no permite modificar el valor de un widget después de que fue creado
    # en el mismo run. Por eso guardamos los valores cargados en claves internas
    # y aumentamos una revisión para recrear los widgets con esos defaults.
    st.session_state.loaded_client_name_v10 = payload.get("client", st.session_state.get("loaded_client_name_v10", "Mercado Livre"))
    st.session_state.loaded_project_name_v10 = payload.get("project", st.session_state.get("loaded_project_name_v10", "Arena Pacaembu"))
    st.session_state.loaded_asking_price_v10 = float(payload.get("asking_price", st.session_state.get("loaded_asking_price_v10", 28.0)) or 0.0)
    st.session_state.loaded_scenario_name_v10 = payload.get("scenario_name", st.session_state.get("loaded_scenario_name_v10", "Base"))
    rows = payload.get("client_input", [])
    st.session_state.client_df_94 = norm_client(pd.DataFrame(rows), mapping) if rows else default_client(mapping)
    st.session_state.form_rev_v10 = int(st.session_state.get("form_rev_v10", 0)) + 1


def read_scenario_file(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def list_saved_scenarios():
    items = []
    for path in sorted(SCENARIO_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            payload = read_scenario_file(path)
            items.append({
                "label": f"{payload.get('client','')} · {payload.get('project','')} · {payload.get('scenario_name', path.stem)}",
                "path": str(path),
                "payload": payload,
                "updated_at": payload.get("updated_at", ""),
            })
        except Exception:
            continue
    return items


def scenario_compare_df(items, master_obj):
    rows = []
    for item in items:
        payload = item.get("payload", {})
        try:
            df = norm_client(pd.DataFrame(payload.get("client_input", [])), master_obj["mapping"])
            res = calc(df, master_obj)
            est = float(res["Valor Estimado USD M"].sum()) if not res.empty else 0.0
            ask = float(payload.get("asking_price", 0) or 0)
            gap = est - ask
            rows.append({
                "Escenario": payload.get("scenario_name", ""),
                "Cliente": payload.get("client", ""),
                "Proyecto": payload.get("project", ""),
                "Valor estimado USD M": round(est, 2),
                "Precio solicitado USD M": round(ask, 2),
                "Gap USD M": round(gap, 2),
                "Gap %": round((gap / ask * 100), 1) if ask else 0.0,
                "Assets": len(df),
                "Actualizado": payload.get("updated_at", ""),
            })
        except Exception:
            pass
    return pd.DataFrame(rows)
# Defaults de UI / escenarios
if "form_rev_v10" not in st.session_state:
    st.session_state.form_rev_v10 = 0
if "loaded_client_name_v10" not in st.session_state:
    st.session_state.loaded_client_name_v10 = "Mercado Livre"
if "loaded_project_name_v10" not in st.session_state:
    st.session_state.loaded_project_name_v10 = "Arena Pacaembu"
if "loaded_asking_price_v10" not in st.session_state:
    st.session_state.loaded_asking_price_v10 = 28.0
if "loaded_scenario_name_v10" not in st.session_state:
    st.session_state.loaded_scenario_name_v10 = "Base"

# Sidebar
# ---------------------------

with st.sidebar:
    show_logo(LOGO_QSPORT, 190)
    st.markdown("### Sponsorship Valuation")
    st.caption("Benchmark Pricing Engine · Scenario Engine · Dual Export · V10.3 Scenario Save/Load Hotfix")
    st.divider()

    access_mode = st.radio("Modo de acceso", ["Cliente", "Interno QSport"], index=0)
    internal_ok = False
    if access_mode == "Interno QSport":
        st.caption("El modo interno habilita benchmark summary y exportables técnicos.")
        pw = st.text_input("Password interno", type="password")
        try:
            secret_pw = st.secrets.get("INTERNAL_PASSWORD", INTERNAL_PASSWORD_DEFAULT)
        except Exception:
            secret_pw = INTERNAL_PASSWORD_DEFAULT
        internal_ok = bool(pw and pw == secret_pw)
        if pw and not internal_ok:
            st.error("Password incorrecto.")
        elif internal_ok:
            st.success("Modo interno habilitado.")
    else:
        st.info("Vista cliente activa: benchmark y export interno ocultos.")

    st.divider()
    form_rev = int(st.session_state.get("form_rev_v10", 0))
    client_name = st.text_input("Cliente", st.session_state.get("loaded_client_name_v10", "Mercado Livre"), key=f"client_name_v10_{form_rev}")
    project_name = st.text_input("Proyecto", st.session_state.get("loaded_project_name_v10", "Arena Pacaembu"), key=f"project_name_v10_{form_rev}")
    asking_price = st.number_input("Precio solicitado por la propiedad (USD M)", min_value=0.0, value=float(st.session_state.get("loaded_asking_price_v10", 28.0)), step=0.5, key=f"asking_price_v10_{form_rev}")

    st.divider()
    uploaded_master = st.file_uploader("Cargar benchmark_latam_master_v1.xlsx", type=["xlsx", "xls"])
    uploaded_client = st.file_uploader("Cargar Excel cliente", type=["xlsx", "xls", "csv"])
    uploaded_scenario_json = st.file_uploader("Cargar escenario guardado (.json)", type=["json"])

    with st.expander("Metodología visible", expanded=False):
        st.write("La herramienta estima valor de mercado con benchmarks comparables y scoring estratégico. No reparte el precio pedido.")

# Guardamos los valores activos como últimos valores conocidos.
st.session_state.loaded_client_name_v10 = client_name
st.session_state.loaded_project_name_v10 = project_name
st.session_state.loaded_asking_price_v10 = float(asking_price or 0.0)

master = load_master(uploaded_master.getvalue() if uploaded_master else None)
if master is None:
    st.error("No se encontró benchmark_latam_master_v1.xlsx. Subilo desde la sidebar o agregalo al repo.")
    st.stop()

if master["deals"].empty:
    st.error("El benchmark cargó, pero Deals_Raw no generó precios válidos. Revisá columnas: Precio Anual, Moneda y FX_Rates.")
    st.stop()

mapping = master["mapping"]

if "client_df_94" not in st.session_state:
    st.session_state.client_df_94 = default_client(mapping)

if uploaded_client:
    try:
        cdf = pd.read_csv(uploaded_client) if uploaded_client.name.lower().endswith(".csv") else pd.read_excel(uploaded_client)
        st.session_state.client_df_94 = norm_client(cdf, mapping)
        st.sidebar.success("Input cliente cargado.")
    except Exception:
        st.sidebar.error("No se pudo leer el input cliente.")


if uploaded_scenario_json:
    try:
        payload = json.loads(uploaded_scenario_json.getvalue().decode("utf-8"))
        load_scenario_payload(payload, mapping)
        st.sidebar.success("Escenario cargado desde archivo JSON.")
    except Exception:
        st.sidebar.error("No se pudo leer el escenario JSON.")

with st.sidebar:
    st.divider()
    st.markdown("### Escenarios")
    st.caption("Guardá, reabrí y compará versiones de una valoración.")
    saved_scenarios = list_saved_scenarios()
    if saved_scenarios:
        labels = [x["label"] for x in saved_scenarios]
        selected_label = st.selectbox("Escenarios guardados", labels, key="saved_scenario_select_v10")
        selected_item = saved_scenarios[labels.index(selected_label)]
        c_load, c_del = st.columns(2)
        with c_load:
            if st.button("Cargar", use_container_width=True):
                load_scenario_payload(selected_item["payload"], mapping)
                st.rerun()
        with c_del:
            if st.button("Eliminar", use_container_width=True):
                try:
                    Path(selected_item["path"]).unlink(missing_ok=True)
                    st.success("Escenario eliminado.")
                    st.rerun()
                except Exception:
                    st.error("No se pudo eliminar.")
    else:
        st.info("Todavía no hay escenarios guardados en esta instalación.")
# Header
mb = b64(LOGO_MELI)
if mb:
    st.markdown(
        f"""
        <div class="qs-card">
          <div style="display:flex;align-items:center;gap:18px;">
            <div style="width:170px;min-width:170px;">
              <img src="data:image/png;base64,{mb}" style="max-width:160px;max-height:58px;object-fit:contain;">
            </div>
            <div>
              <div class="qs-title">Sponsorship / Naming Rights Valuation Tool</div>
              <div class="qs-subtitle">Independent valuation of sponsorship assets vs asking price</div>
              <div style="margin-top:10px;">
                <span class="qs-pill">Prepared for {client_name}</span>
                <span class="qs-pill">{access_mode}</span>
                <span class="qs-pill">Confidential</span>
              </div>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    st.markdown('<div class="qs-card"><div class="qs-title">Sponsorship / Naming Rights Valuation Tool</div><div class="qs-subtitle">Independent valuation of sponsorship assets vs asking price</div></div>', unsafe_allow_html=True)


st.markdown('<div class="section-title">1. Assets a valorar</div>', unsafe_allow_html=True)
st.markdown('<div class="small-muted">El cliente informa assets, país y marca. QSport completa la evaluación estratégica en la app.</div>', unsafe_allow_html=True)

visible_types = sorted(mapping["Tipo Visible Cliente"].dropna().astype(str).unique().tolist())
countries = sorted(set(["LATAM", "AR", "BR", "CL", "CO", "MX", "PE", "UY", "EC"] + master["deals"]["País"].dropna().astype(str).unique().tolist()))

edited = st.data_editor(
    st.session_state.client_df_94,
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True,
    column_config={
        "Asset": st.column_config.TextColumn("Asset", width="large"),
        "Tipo Visible Cliente": st.column_config.SelectboxColumn("Tipo Visible Cliente", options=visible_types, required=True),
        "País": st.column_config.SelectboxColumn("País", options=countries, required=True),
        "Marca Interesada": st.column_config.TextColumn("Marca Interesada"),
        **{c: st.column_config.SelectboxColumn(c, options=list(SCORE_LABELS.keys()), required=True) for c in SCORE_COLS},
    },
    key=f"editor_94_dual_{int(st.session_state.get('form_rev_v10', 0))}",
)

st.session_state.client_df_94 = norm_client(edited, mapping)

st.markdown('<div class="section-title">2. Scenario Engine</div>', unsafe_allow_html=True)
st.markdown('<div class="small-muted">Guardá la valoración como escenario, descargala como archivo JSON o duplicala para probar otra versión sin perder la anterior.</div>', unsafe_allow_html=True)
form_rev = int(st.session_state.get("form_rev_v10", 0))
scenario_name = st.text_input("Nombre del escenario", st.session_state.get("loaded_scenario_name_v10", "Base"), key=f"scenario_name_v10_{form_rev}")
st.session_state.loaded_scenario_name_v10 = scenario_name
sc1, sc2, sc3, sc4 = st.columns(4)
with sc1:
    if st.button("Guardar escenario", use_container_width=True):
        path = save_scenario_local(client_name, project_name, asking_price, scenario_name, st.session_state.client_df_94)
        st.success(f"Escenario guardado: {path.name}")
with sc2:
    if st.button("Duplicar escenario", use_container_width=True):
        copy_name = f"{scenario_name} copia"
        path = save_scenario_local(client_name, project_name, asking_price, copy_name, st.session_state.client_df_94)
        st.session_state.loaded_scenario_name_v10 = copy_name
        st.session_state.form_rev_v10 = int(st.session_state.get("form_rev_v10", 0)) + 1
        st.success(f"Duplicado creado: {path.name}")
        st.rerun()
with sc3:
    st.download_button(
        "Descargar escenario JSON",
        data=scenario_to_bytes(client_name, project_name, asking_price, scenario_name, st.session_state.client_df_94),
        file_name=f"{_safe_filename(client_name)}_{_safe_filename(project_name)}_{_safe_filename(scenario_name)}_scenario_v10.json",
        mime="application/json",
        use_container_width=True,
    )
with sc4:
    st.download_button(
        "Descargar input Excel",
        data=build_reloadable_input_excel(st.session_state.client_df_94),
        file_name=f"{_safe_filename(client_name)}_{_safe_filename(project_name)}_{_safe_filename(scenario_name)}_input_v10.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
result = calc(st.session_state.client_df_94, master)

estimated_value = float(result["Valor Estimado USD M"].sum()) if not result.empty else 0.0
gap = estimated_value - asking_price
gap_pct = (gap / asking_price * 100) if asking_price else 0.0
assets_count = len(result)
comparables = int(result["Comparables"].sum()) if not result.empty else 0

if asking_price > 0:
    if gap_pct >= 10:
        rec = "Opportunity: estimated value is above asking price."
        rec_es = "El paquete presenta una oportunidad frente al precio solicitado."
    elif gap_pct <= -10:
        rec = "Overpriced: asking price is above estimated benchmark value."
        rec_es = "El paquete presenta sobreprecio frente al benchmark estimado."
    else:
        rec = "Fair range: asking price is aligned with estimated benchmark value."
        rec_es = "El precio solicitado se encuentra en zona razonable."
else:
    rec = "Pending asking price."
    rec_es = "Ingresá el precio solicitado para comparar."

st.markdown('<div class="section-title">3. Resultado ejecutivo</div>', unsafe_allow_html=True)
for col, label, value, note in zip(
    st.columns(4),
    ["Valor estimado", "Precio solicitado", "Gap vs precio", "Assets / comparables"],
    [money(estimated_value), money(asking_price), money(gap), f"{assets_count} / {comparables}"],
    ["Benchmark + scoring", "Oferta propiedad", f"{gap_pct:.1f}%", "Assets valorados y casos usados"],
):
    with col:
        st.markdown(f'<div class="metric-box"><div class="metric-label">{label}</div><div class="metric-value">{value}</div><div class="metric-note">{note}</div></div>', unsafe_allow_html=True)

st.markdown(f'<div class="qs-card"><b>Lectura estratégica</b><br><span style="color:#475467;">{rec_es}</span></div>', unsafe_allow_html=True)

st.markdown('<div class="section-title">4. Valuación por asset</div>', unsafe_allow_html=True)
display = result.sort_values("Valor Estimado USD M", ascending=False).reset_index(drop=True)

if access_mode == "Cliente" or not internal_ok:
    client_display_cols = ["Asset", "Tipo Visible Cliente", "Score Final", "Clasificación", "Valor Estimado USD M", "Rango Bajo Ajustado USD M", "Rango Alto Ajustado USD M"]
    st.dataframe(display[client_display_cols] if not display.empty else display, use_container_width=True, hide_index=True)
else:
    st.dataframe(display, use_container_width=True, hide_index=True)

c1, c2 = st.columns(2)
with c1:
    st.markdown('<div class="section-title">Valor estimado por asset</div>', unsafe_allow_html=True)
    if not display.empty:
        fig = px.bar(display.head(15).sort_values("Valor Estimado USD M"), x="Valor Estimado USD M", y="Asset", orientation="h", text="Valor Estimado USD M")
        fig.update_layout(height=470, margin=dict(l=10, r=10, t=20, b=10), xaxis_title="USD M", yaxis_title="", plot_bgcolor="white", paper_bgcolor="white")
        st.plotly_chart(fig, use_container_width=True)

with c2:
    st.markdown('<div class="section-title">Distribución por tipo visible</div>', unsafe_allow_html=True)
    if not display.empty:
        type_df = display.groupby("Tipo Visible Cliente", as_index=False)["Valor Estimado USD M"].sum().sort_values("Valor Estimado USD M", ascending=False)
        fig2 = px.pie(type_df, values="Valor Estimado USD M", names="Tipo Visible Cliente", hole=.55)
        fig2.update_layout(height=470, margin=dict(l=10, r=10, t=20, b=10), paper_bgcolor="white")
        st.plotly_chart(fig2, use_container_width=True)

st.markdown('<div class="section-title">5. Comparador de escenarios</div>', unsafe_allow_html=True)
saved_for_compare = list_saved_scenarios()
if saved_for_compare:
    labels_compare = [x["label"] for x in saved_for_compare]
    selected_compare = st.multiselect("Seleccioná escenarios para comparar", labels_compare, default=labels_compare[:min(3, len(labels_compare))])
    if selected_compare:
        selected_items = [saved_for_compare[labels_compare.index(lbl)] for lbl in selected_compare]
        cmp_df = scenario_compare_df(selected_items, master)
        if not cmp_df.empty:
            st.dataframe(cmp_df, use_container_width=True, hide_index=True)
            fig_cmp = px.bar(cmp_df.sort_values("Valor estimado USD M"), x="Valor estimado USD M", y="Escenario", orientation="h", text="Valor estimado USD M")
            fig_cmp.update_layout(height=330, margin=dict(l=10, r=10, t=20, b=10), xaxis_title="USD M", yaxis_title="", plot_bgcolor="white", paper_bgcolor="white")
            st.plotly_chart(fig_cmp, use_container_width=True)
else:
    st.info("Guardá al menos un escenario para activar la comparación.")

bench_summary = master["deals"].groupby("Familia Benchmark", as_index=False).agg(
    Casos=("Precio USD Anual M", "count"),
    Mediana_USD_M=("Precio USD Anual M", "median"),
    Promedio_USD_M=("Precio USD Anual M", "mean"),
    Min_USD_M=("Precio USD Anual M", "min"),
    Max_USD_M=("Precio USD Anual M", "max"),
).round(2).sort_values("Casos", ascending=False)

if access_mode == "Interno QSport" and internal_ok:
    st.markdown('<div class="section-title">6. Benchmark summary interno</div>', unsafe_allow_html=True)
    st.dataframe(bench_summary, use_container_width=True, hide_index=True)
else:
    st.markdown('<div class="section-title">6. Benchmark summary</div>', unsafe_allow_html=True)
    st.info("Resumen interno oculto en modo cliente.")

summary = {
    "client": client_name,
    "project": project_name,
    "asking_price": asking_price,
    "estimated_value": estimated_value,
    "gap": gap,
    "gap_pct": gap_pct,
    "recommendation": rec,
}

st.markdown('<div class="section-title">7. Exportables</div>', unsafe_allow_html=True)

try:
    client_pdf_bytes = build_client_pdf(display, summary, str(LOGO_QSPORT), str(LOGO_MELI))
    pdf_error = None
except Exception as e:
    client_pdf_bytes = None
    pdf_error = str(e)

client_excel_bytes = build_client_excel(display, summary)
internal_excel_bytes = build_excel(display, st.session_state.client_df_94, summary, bench_summary)

st.markdown('<div class="qs-card"><b>Export Cliente</b><br><span style="color:#475467;">Outputs limpios para compartir: sin benchmark interno, sin factores propietarios y sin detalle sensible de comparables.</span></div>', unsafe_allow_html=True)
ce1, ce2 = st.columns(2)
with ce1:
    if client_pdf_bytes:
        st.download_button("Descargar PDF Cliente", data=client_pdf_bytes, file_name=f"{_safe_filename(client_name)}_{_safe_filename(project_name)}_client_report_v10.pdf", mime="application/pdf", use_container_width=True)
    else:
        st.error("No se pudo generar el PDF cliente. El Excel cliente sigue disponible.")
        if access_mode == "Interno QSport" and internal_ok:
            with st.expander("Ver detalle técnico"):
                st.code(pdf_error or "Error desconocido")
with ce2:
    st.download_button("Descargar Excel Cliente Resumido", data=client_excel_bytes, file_name=f"{_safe_filename(client_name)}_{_safe_filename(project_name)}_client_summary_v10.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

if access_mode == "Interno QSport" and internal_ok:
    st.markdown('<div class="qs-card"><b>Export Interno QSport</b><br><span style="color:#475467;">Outputs técnicos completos para auditoría, revisión comercial y continuidad interna.</span></div>', unsafe_allow_html=True)
    ie1, ie2 = st.columns(2)
    with ie1:
        st.download_button("Descargar Excel Interno Completo", data=internal_excel_bytes, file_name=f"{_safe_filename(client_name)}_{_safe_filename(project_name)}_internal_full_v10.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with ie2:
        if client_pdf_bytes:
            st.download_button("Descargar PDF Interno Técnico", data=client_pdf_bytes, file_name=f"{_safe_filename(client_name)}_{_safe_filename(project_name)}_internal_report_v10.pdf", mime="application/pdf", use_container_width=True)
else:
    st.caption("Los exportables internos solo aparecen con password QSport.")

st.markdown('<div style="margin-top:24px;padding-top:14px;border-top:1px solid #e6e8ef;color:#98a2b3;font-size:12px;">Confidential valuation output. Benchmark database is proprietary and not disclosed in full to external users.</div>', unsafe_allow_html=True)
